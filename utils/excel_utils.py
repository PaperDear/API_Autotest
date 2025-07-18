import os

import openpyxl

from config.config import CASE_FILE, SHEET_NAME


def read_excel(case_file=CASE_FILE,sheet_name=SHEET_NAME):
    # 本地测试本文件执行时打开
    wb = openpyxl.load_workbook(case_file)

    # pytest执行时打开
    # wb = openpyxl.load_workbook(os.getcwd() + '/data/cases.xlsx')
    # 选择表sheet1
    sheet = wb[sheet_name]
    # 读数据
    data = []
    # 首行为中文注解，提取key从第二行的英文开始
    keys = [cell.value for cell in sheet[2]]
    # value从第3行开始提取
    for row in sheet.iter_rows(min_row=3, values_only=True):
        dict_data = dict(zip(keys, row))
        # 判断是否是需要执行的用例
        if dict_data['isTrue']:
            data.append(dict_data)
    print(data)
    # 关闭文件
    wb.close()
    return data

# if __name__ == '__main__':
#     read_excel()
